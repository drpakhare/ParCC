#!/usr/bin/env python3
"""
Create ParCC v1.4 Validation Test Cases Workbook with openpyxl.

This script creates a comprehensive Excel workbook for validating ParCC v1.4 modules,
following the exact structure and format of the v1.3 version.
"""

import math
import copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter

# ============================================================================
# STYLE DEFINITIONS
# ============================================================================

# Colors
DARK_BLUE = "FF003366"
LIGHT_GRAY = "FFF2F2F2"
LIGHT_BLUE = "FFD5E8F0"
BLUE_FONT = "FF0000FF"
GREEN_FONT = "FF008000"
ORANGE_FILL = "FFFFA500"
LIGHT_GREEN_FILL = "FFE2EFDA"
LIGHT_BLUE_FILL = "FFDAEEF3"

# Fonts
HEADER_FONT = Font(name='Arial', size=11, bold=True, color=DARK_BLUE)
TEST_HEADER_FONT = Font(name='Arial', size=12, bold=True, color=DARK_BLUE)
NORMAL_FONT = Font(name='Arial', size=11)
BLUE_INPUT_FONT = Font(name='Arial', size=11, color=BLUE_FONT)
GREEN_FORMULA_FONT = Font(name='Arial', size=11, color=GREEN_FONT)

# Fills
LIGHT_GRAY_FILL = PatternFill(fill_type='solid', fgColor=LIGHT_GRAY)
LIGHT_BLUE_HEADER_FILL = PatternFill(fill_type='solid', fgColor=LIGHT_BLUE)
GREEN_FORMULA_FILL = PatternFill(fill_type='solid', fgColor=LIGHT_GREEN_FILL)
BLUE_INPUT_FILL = PatternFill(fill_type='solid', fgColor=LIGHT_BLUE_FILL)
ORANGE_KEY_FILL = PatternFill(fill_type='solid', fgColor=ORANGE_FILL)

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def setup_sheet_columns(ws):
    """Set up column widths for a test sheet."""
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    for col in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15

def set_cell_style(cell, font=None, fill=None, alignment=None):
    """Apply styling to a cell."""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment

def create_test_header(ws, row, test_name):
    """Create a test header with merged cells across A-F."""
    ws.merge_cells(f'A{row}:F{row}')
    cell = ws[f'A{row}']
    cell.value = test_name
    set_cell_style(cell, font=TEST_HEADER_FONT, fill=LIGHT_GRAY_FILL)
    return row + 1

def add_input_row(ws, row, label, value, source_col='C', source_text=''):
    """Add an input row (label, blue input, source info)."""
    ws[f'A{row}'].value = label
    ws[f'B{row}'].value = value
    if source_text:
        ws[f'{source_col}{row}'].value = source_text

    set_cell_style(ws[f'A{row}'], fill=LIGHT_GRAY_FILL)
    set_cell_style(ws[f'B{row}'], font=BLUE_INPUT_FONT, fill=BLUE_INPUT_FILL)
    if source_text:
        set_cell_style(ws[f'{source_col}{row}'], fill=LIGHT_GRAY_FILL)

    return row + 1

def add_formula_row(ws, row, label, formula):
    """Add a formula row (label, green formula)."""
    ws[f'A{row}'].value = label
    ws[f'B{row}'].value = formula

    set_cell_style(ws[f'A{row}'], fill=LIGHT_GRAY_FILL)
    set_cell_style(ws[f'B{row}'], font=GREEN_FORMULA_FONT, fill=GREEN_FORMULA_FILL)

    return row + 1

def add_parcc_output_row(ws, row, label, expected_row):
    """Add ParCC output comparison row (white input cell, match formula)."""
    ws[f'A{row}'].value = label
    ws[f'C{row}'].value = 'Match?'

    # Match formula: check if ParCC output matches expected (empty = PENDING)
    ws[f'D{row}'].value = f'=IF(B{row}="","PENDING",IF(ROUND(B{row},5)=ROUND(B{expected_row},5),"PASS","FAIL"))'

    set_cell_style(ws[f'A{row}'], fill=LIGHT_GRAY_FILL)
    set_cell_style(ws[f'C{row}'], fill=LIGHT_GRAY_FILL)
    set_cell_style(ws[f'D{row}'], font=GREEN_FORMULA_FONT, fill=GREEN_FORMULA_FILL)

    return row + 1

def add_blank_row(ws, row):
    """Add a blank separator row."""
    return row + 1

# ============================================================================
# CREATE WORKBOOK
# ============================================================================

def create_workbook():
    """Build the complete ParCC v1.4 validation workbook."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # ========================================================================
    # INSTRUCTIONS SHEET
    # ========================================================================
    ws = wb.create_sheet('Instructions', 0)
    setup_sheet_columns(ws)
    row = 1

    ws[f'A{row}'].value = 'PARCC VALIDATION TEST CASES'
    set_cell_style(ws[f'A{row}'], font=HEADER_FONT)
    row += 2

    ws[f'A{row}'].value = 'PURPOSE'
    set_cell_style(ws[f'A{row}'], font=HEADER_FONT)
    row += 1

    ws[f'A{row}'].value = 'This workbook validates every ParCC module using pre-computed test cases with Excel formulas.'
    row += 2

    ws[f'A{row}'].value = 'WORKBOOK STRUCTURE'
    set_cell_style(ws[f'A{row}'], font=HEADER_FONT)
    row += 1

    ws[f'A{row}'].value = '  Each sheet has two sections:'
    row += 1

    ws[f'A{row}'].value = '  (A) EXCEL CALCULATOR — Input cells (blue) and formula cells (green) that compute the expected result'
    row += 1

    ws[f'A{row}'].value = '  (B) PARCC COMPARISON — Enter the ParCC output; a formula compares against the Excel result'
    row += 2

    ws[f'A{row}'].value = 'TOTAL: 51 test cases across 15 module sheets'
    row += 2

    ws[f'A{row}'].value = 'COLOR LEGEND'
    set_cell_style(ws[f'A{row}'], font=HEADER_FONT)
    row += 1

    ws[f'A{row}'].value = 'Blue text'
    ws[f'B{row}'].value = '= Input values'
    set_cell_style(ws[f'B{row}'], font=BLUE_INPUT_FONT)
    row += 1

    ws[f'A{row}'].value = 'Green text'
    ws[f'B{row}'].value = '= Excel formulas (expected output)'
    set_cell_style(ws[f'B{row}'], font=GREEN_FORMULA_FONT)
    row += 1

    ws[f'A{row}'].value = 'Orange fill'
    ws[f'B{row}'].value = '= Key results'
    set_cell_style(ws[f'B{row}'], fill=ORANGE_KEY_FILL)
    row += 1

    ws[f'A{row}'].value = 'White cells'
    ws[f'B{row}'].value = '= ParCC output (to be filled by validator)'

    # ========================================================================
    # SUMMARY SHEET
    # ========================================================================
    ws = wb.create_sheet('Summary', 1)
    setup_sheet_columns(ws)
    row = 1

    ws[f'A{row}'].value = 'Module'
    ws[f'B{row}'].value = 'Tests'
    ws[f'C{row}'].value = 'PASS'
    for col in ['A', 'B', 'C']:
        set_cell_style(ws[f'{col}{row}'], font=Font(name='Arial', size=11, bold=True),
                      fill=LIGHT_BLUE_HEADER_FILL)
    row += 1

    # Module list with test counts
    modules = [
        ('1. Core Converters', 6, ['D6', 'D14', 'D19', 'D24', 'D31', 'D37']),
        ('2. HR Converter', 4, ['D11', 'D20', 'D29', 'D38']),
        ('3. Survival', 3, ['D7', 'D14', 'D27']),
        ('4. Bg Mortality', 5, ['D8', 'D16', 'D26', 'D36', 'D44']),
        ('5. PSA Distributions', 4, ['D8', 'D17', 'D24', 'D32']),
        ('6. Financial', 4, ['D8', 'D16', 'D24', 'D31']),
        ('7. Diagnostics', 3, ['D7', 'D15', 'D23']),
        ('8. ICER & NMB', 3, ['D8', 'D14', 'D21']),
        ('9. VBP', 3, ['D6', 'D16', 'D21']),
        ('10. OR-RR & Effect Size', 3, ['D7', 'D17', 'D28']),
        ('11. NNT & Log-rank HR', 3, ['D8', 'D16', 'D24']),
        ('12. Dirichlet', 2, ['D9', 'D18']),
        ('13. Log-Logistic Survival', 2, ['D15', 'D27']),
        ('14. Budget Impact Analysis', 2, ['D10', 'D21']),
        ('15. PPP Converter', 3, ['D14', 'D23', 'D32']),
    ]

    for module_name, test_count, match_cells in modules:
        ws[f'A{row}'].value = module_name
        ws[f'B{row}'].value = test_count
        # Sum PASS counts from all tests in this module
        formula = '+'.join([f'COUNTIF(\'{module_name}\'!{cell},"PASS")' for cell in match_cells])
        ws[f'C{row}'].value = f'={formula}'
        row += 1

    ws[f'A{row}'].value = 'TOTAL'
    ws[f'B{row}'].value = '=SUM(B2:B16)'
    ws[f'C{row}'].value = '=SUM(C2:C16)'
    for col in ['A', 'B', 'C']:
        set_cell_style(ws[f'{col}{row}'], font=Font(name='Arial', size=11, bold=True))
    row += 1

    ws[f'A{row}'].value = 'OVERALL RESULT'
    ws[f'B{row}'].value = '=IF(D16=0,IF(C16=B16,"ALL PASS","PENDING"),"FAILURES FOUND")'
    set_cell_style(ws[f'A{row}'], font=Font(name='Arial', size=11, bold=True))

    # ========================================================================
    # COPY EXISTING SHEETS (v1.3)
    # ========================================================================
    from openpyxl import load_workbook as load_wb

    existing_wb = load_wb('/sessions/trusting-practical-tesla/mnt/ParCC/ParCC_Validation_TestCases.xlsx')

    for sheet_name in ['1. Core Converters', '2. HR Converter', '3. Survival', '4. Bg Mortality',
                       '5. PSA Distributions', '6. Financial', '7. Diagnostics', '8. ICER & NMB', '9. VBP']:
        source_ws = existing_wb[sheet_name]
        target_ws = wb.create_sheet(sheet_name)

        # Copy all cells - copy values and use deep copy for styles
        for row in source_ws.iter_rows():
            for cell in row:
                target_cell = target_ws[cell.coordinate]
                target_cell.value = cell.value

                # Deep copy font and fill using copy module
                if cell.font:
                    try:
                        target_cell.font = copy.copy(cell.font)
                    except:
                        pass

                if cell.fill and cell.fill.fill_type and cell.fill.fill_type != 'none':
                    try:
                        target_cell.fill = copy.copy(cell.fill)
                    except:
                        pass

        # Copy column widths
        for col in source_ws.column_dimensions:
            if source_ws.column_dimensions[col].width:
                target_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width

    # ========================================================================
    # NEW SHEET 10: OR-RR & Effect Size
    # ========================================================================
    ws = wb.create_sheet('10. OR-RR & Effect Size')
    setup_sheet_columns(ws)
    row = 1

    # TEST 1: OR to RR (Zhang & Yu)
    row = create_test_header(ws, row, 'TEST 1: OR to RR (Zhang & Yu) — Depression NMA')
    row = add_input_row(ws, row, 'OR', 1.85, 'C', 'Source:')
    row = add_input_row(ws, row, 'Baseline risk (p0)', 0.30, 'D', 'Depression prevalence')
    row = add_formula_row(ws, row, 'Excel: RR = OR / (1 - p0 + p0*OR)', '=B3/(1-B4+B4*B3)')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)
    row = add_blank_row(ws, row)

    # TEST 2: RR to OR (reverse)
    row = create_test_header(ws, row, 'TEST 2: RR to OR (reverse)')
    row = add_input_row(ws, row, 'RR', 1.50)
    row = add_input_row(ws, row, 'Baseline risk (p0)', 0.20)
    row = add_formula_row(ws, row, 'Excel: OR = RR * (1-p0) / (1 - RR*p0)', '=B10*(1-B11)/(1-B10*B11)')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)
    row = add_blank_row(ws, row)

    # TEST 3: SMD to log(OR) (Chinn 2000)
    row = create_test_header(ws, row, 'TEST 3: SMD to log(OR) (Chinn 2000)')
    row = add_input_row(ws, row, 'SMD', 0.50)
    row = add_formula_row(ws, row, 'log(OR) = SMD * π / √3', '=B18*PI()/SQRT(3)')
    row = add_formula_row(ws, row, 'OR = EXP(log(OR))', '=EXP(B19)')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)

    # ========================================================================
    # NEW SHEET 11: NNT & Log-rank HR
    # ========================================================================
    ws = wb.create_sheet('11. NNT & Log-rank HR')
    setup_sheet_columns(ws)
    row = 1

    # TEST 1: NNT from ARR
    row = create_test_header(ws, row, 'TEST 1: NNT from ARR')
    row = add_input_row(ws, row, 'Control probability', 0.15)
    row = add_input_row(ws, row, 'Intervention probability', 0.10)
    row = add_formula_row(ws, row, 'ARR = p_control - p_int', '=B2-B3')
    row = add_formula_row(ws, row, 'NNT = CEILING(1/ARR, 1)', '=CEILING(1/B4,1)')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)
    row = add_blank_row(ws, row)

    # TEST 2: NNT from RR + baseline
    row = create_test_header(ws, row, 'TEST 2: NNT from RR + baseline')
    row = add_input_row(ws, row, 'RR', 0.75)
    row = add_input_row(ws, row, 'Baseline risk (p0)', 0.20)
    row = add_formula_row(ws, row, 'ARR = p0 * (1 - RR)', '=B10*(1-B9)')
    row = add_formula_row(ws, row, 'NNT = CEILING(1/ARR, 1)', '=CEILING(1/B11,1)')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)
    row = add_blank_row(ws, row)

    # TEST 3: Log-rank to HR (Peto approximation)
    row = create_test_header(ws, row, 'TEST 3: Log-rank to HR (Peto approximation)')
    row = add_input_row(ws, row, 'Chi-square', 6.25)
    row = add_input_row(ws, row, 'Total events', 100)
    row = add_formula_row(ws, row, 'z = SQRT(Chi-square)', '=SQRT(B17)')
    row = add_formula_row(ws, row, 'log(HR) = z / SQRT(events/4)', '=B18/SQRT(B19/4)')
    row = add_formula_row(ws, row, 'HR = EXP(log(HR))', '=EXP(B20)')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)

    # ========================================================================
    # NEW SHEET 12: Dirichlet
    # ========================================================================
    ws = wb.create_sheet('12. Dirichlet')
    setup_sheet_columns(ws)
    row = 1

    # TEST 1: Dirichlet parameters from transition counts
    row = create_test_header(ws, row, 'TEST 1: Dirichlet parameters from transition counts')
    row = add_input_row(ws, row, 'Stable count', 150)
    row = add_input_row(ws, row, 'Progressed count', 35)
    row = add_input_row(ws, row, 'Dead count', 15)
    row = add_formula_row(ws, row, 'Total observations', '=B2+B3+B4')
    row = add_formula_row(ws, row, 'Expected: Stable proportion', '=B2/B5')
    row = add_formula_row(ws, row, 'Expected: Progressed proportion', '=B3/B5')
    row = add_formula_row(ws, row, 'Expected: Dead proportion', '=B4/B5')
    row = add_formula_row(ws, row, 'Sum of proportions', '=B6+B7+B8')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)
    row = add_blank_row(ws, row)

    # TEST 2: Dirichlet concentration check
    row = create_test_header(ws, row, 'TEST 2: Dirichlet concentration check')
    row = add_input_row(ws, row, 'A count', 80)
    row = add_input_row(ws, row, 'B count', 15)
    row = add_input_row(ws, row, 'C count', 5)
    row = add_formula_row(ws, row, 'Total observations', '=B11+B12+B13')
    row = add_formula_row(ws, row, 'Expected: A proportion', '=B11/B14')
    row = add_formula_row(ws, row, 'Expected: B proportion', '=B12/B14')
    row = add_formula_row(ws, row, 'Expected: C proportion', '=B13/B14')
    row = add_formula_row(ws, row, 'Sum of proportions', '=B15+B16+B17')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)

    # ========================================================================
    # NEW SHEET 13: Log-Logistic Survival
    # ========================================================================
    ws = wb.create_sheet('13. Log-Logistic Survival')
    setup_sheet_columns(ws)
    row = 1

    # TEST 1: Two-point Log-Logistic fit
    row = create_test_header(ws, row, 'TEST 1: Two-point Log-Logistic fit')
    row = add_input_row(ws, row, 'S(t1) at t1=12 months', 0.80)
    row = add_input_row(ws, row, 't1 (months)', 12)
    row = add_input_row(ws, row, 'S(t2) at t2=36 months', 0.40)
    row = add_input_row(ws, row, 't2 (months)', 36)
    row = add_formula_row(ws, row, 'log-odds at t1: log((1-S)/S)', '=LN((1-B2)/B2)')
    row = add_formula_row(ws, row, 'log-odds at t2: log((1-S)/S)', '=LN((1-B4)/B4)')
    row = add_formula_row(ws, row, 'β = (logodds2 - logodds1) / (ln(t2) - ln(t1))', '=(B7-B6)/(LN(B5)-LN(B3))')
    row = add_formula_row(ws, row, 'ln(α) = [ln(t1)*logodds2 - ln(t2)*logodds1] / (logodds2 - logodds1)', '=(LN(B3)*B7-LN(B5)*B6)/(B7-B6)')
    row = add_formula_row(ws, row, 'α = EXP(ln(α))', '=EXP(B9)')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)
    row = add_blank_row(ws, row)

    # TEST 2: Verify survival at third timepoint
    row = create_test_header(ws, row, 'TEST 2: Predict S(24) using fitted α and β')
    row = add_input_row(ws, row, 'Fitted α from TEST 1', '=B10')
    row = add_input_row(ws, row, 'Fitted β from TEST 1', '=B8')
    row = add_input_row(ws, row, 'Prediction time t (months)', 24)
    row = add_formula_row(ws, row, 'S(t) = 1 / (1 + (t/α)^β)', '=1/(1+(B14/B12)^B13)')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)

    # ========================================================================
    # NEW SHEET 14: Budget Impact Analysis
    # ========================================================================
    ws = wb.create_sheet('14. Budget Impact Analysis')
    setup_sheet_columns(ws)
    row = 1

    # TEST 1: Simple BIA (Year 1, no discounting)
    row = create_test_header(ws, row, 'TEST 1: Simple BIA (Year 1 only)')
    row = add_input_row(ws, row, 'Population', 100000)
    row = add_input_row(ws, row, 'Prevalence', 0.01)
    row = add_input_row(ws, row, 'Proportion eligible', 0.60)
    row = add_input_row(ws, row, 'Uptake Year 1', 0.10)
    row = add_input_row(ws, row, 'Current treatment cost', 8000)
    row = add_input_row(ws, row, 'New treatment cost', 22000)
    row = add_formula_row(ws, row, 'N eligible = Pop * Prev * Eligible', '=B2*B3*B4')
    row = add_formula_row(ws, row, 'N adopters = N eligible * Uptake', '=B8*B5')
    row = add_formula_row(ws, row, 'Incremental cost per patient', '=B7-B6')
    row = add_formula_row(ws, row, 'Budget Impact = N adopters * Inc cost', '=B9*B10')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)
    row = add_blank_row(ws, row)

    # TEST 2: BIA with discounting (Year 3)
    row = create_test_header(ws, row, 'TEST 2: BIA with discounting (Year 3)')
    row = add_input_row(ws, row, 'Population', 100000)
    row = add_input_row(ws, row, 'Prevalence', 0.01)
    row = add_input_row(ws, row, 'Proportion eligible', 0.60)
    row = add_input_row(ws, row, 'Uptake Year 3', 0.30)
    row = add_input_row(ws, row, 'Discount rate', 0.03)
    row = add_input_row(ws, row, 'Current treatment cost', 8000)
    row = add_input_row(ws, row, 'New treatment cost', 22000)
    row = add_formula_row(ws, row, 'N eligible = Pop * Prev * Eligible', '=B15*B16*B17')
    row = add_formula_row(ws, row, 'N adopters = N eligible * Uptake', '=B21*B18')
    row = add_formula_row(ws, row, 'Incremental cost per patient', '=B20-B19')
    row = add_formula_row(ws, row, 'Undiscounted BI = N adopters * Inc cost', '=B22*B23')
    row = add_formula_row(ws, row, 'Discounted BI = Undiscounted / (1+rate)^2', '=B24/(1+B18)^2')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)

    # ========================================================================
    # NEW SHEET 15: PPP Converter
    # ========================================================================
    ws = wb.create_sheet('15. PPP Converter')
    setup_sheet_columns(ws)
    row = 1

    # TEST 1: US to India PPP conversion
    row = create_test_header(ws, row, 'TEST 1: US to India PPP conversion')
    row = add_input_row(ws, row, 'Source cost (USD)', 100000)
    row = add_input_row(ws, row, 'PPP factor US', 1.00, 'C', 'Reference country')
    row = add_input_row(ws, row, 'PPP factor India', 22.88, 'C', 'World Bank 2022 ICP')
    row = add_input_row(ws, row, 'Market FX rate (INR/USD)', 82.78)
    row = add_formula_row(ws, row, 'International$ = Cost / PPP_source', '=B2/B3')
    row = add_formula_row(ws, row, 'Target LCU (PPP-adjusted)', '=B6*B4')
    row = add_formula_row(ws, row, 'Target LCU (Market FX)', '=B2*B5')
    row = add_formula_row(ws, row, 'PPP/FX ratio', '=B7/B8')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)
    row = add_blank_row(ws, row)

    # TEST 2: WHO-CHOICE WTP threshold
    row = create_test_header(ws, row, 'TEST 2: WHO-CHOICE WTP threshold')
    row = add_input_row(ws, row, 'GDP per capita India (USD)', 2389)
    row = add_input_row(ws, row, 'Market FX rate (INR/USD)', 82.78)
    row = add_formula_row(ws, row, 'GDP per capita (INR)', '=B13*B14')
    row = add_formula_row(ws, row, '1x GDP threshold (INR)', '=B15')
    row = add_formula_row(ws, row, '3x GDP threshold (INR)', '=3*B15')
    row = add_input_row(ws, row, 'PPP-converted cost (INR)', 2288000)
    row = add_formula_row(ws, row, 'Is cost > 3x threshold?', '=IF(B20>B17,"YES","NO")')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)
    row = add_blank_row(ws, row)

    # TEST 3: UK to India PPP conversion
    row = create_test_header(ws, row, 'TEST 3: UK to India PPP conversion')
    row = add_input_row(ws, row, 'Source cost (GBP)', 30000)
    row = add_input_row(ws, row, 'PPP factor UK', 0.69)
    row = add_input_row(ws, row, 'PPP factor India', 22.88)
    row = add_formula_row(ws, row, 'International$ = Cost / PPP_source', '=B25/B26')
    row = add_formula_row(ws, row, 'Target LCU (INR, PPP-adjusted)', '=B29*B27')
    row = add_parcc_output_row(ws, row, 'ParCC Output', row - 1)

    return wb

# ============================================================================
# MAIN
# ============================================================================

if __name__ == '__main__':
    print("Creating ParCC v1.4 Validation Test Cases workbook...")
    wb = create_workbook()

    output_path = '/sessions/trusting-practical-tesla/mnt/ParCC/ParCC_Validation_TestCases.xlsx'
    wb.save(output_path)
    print(f"Workbook saved to {output_path}")
    print("\nSheet summary:")
    for sheet_name in wb.sheetnames:
        print(f"  - {sheet_name}")
